import csv
import re
from urllib.parse import urlencode

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth import get_user_model
from django.core.paginator import Paginator
from django.db.models import Count, Prefetch, Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils.dateparse import parse_date
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font

from .models import (
    AuditLog,
    BajaEquipo,
    CentroCosto,
    Equipo,
    ImportLog,
    Marca,
    ModeloEquipo,
    MotivoBaja,
    SistemaOperativo,
    TipoEquipo,
)
from .permissions import (
    can_audit,
    can_baja,
    can_edit,
    can_import,
    can_view_report,
)


IP_REGEX = re.compile(r"^(\d{1,3}\.){3}\d{1,3}$")
MAC_REGEX = re.compile(r"^([0-9A-Fa-f]{2}[:-]){5}([0-9A-Fa-f]{2})$")


def _build_querystring(request, exclude=None, extra=None):
    params = request.GET.copy()
    if exclude:
        for key in exclude:
            params.pop(key, None)
    if extra:
        for key, value in extra.items():
            params[key] = value
    return params.urlencode()


@login_required
def equipos_list(request):
    texto = request.GET.get("texto", "").strip()
    sociedad_id = request.GET.get("sociedad")
    division_id = request.GET.get("division")
    centro_costo_id = request.GET.get("centro_costo")
    marca_id = request.GET.get("marca")
    sistema_operativo_id = request.GET.get("sistema_operativo")
    tipo_equipo_id = request.GET.get("tipo_equipo")
    entidad = request.GET.get("entidad", "").strip()
    municipio = request.GET.get("municipio", "").strip()
    estado = request.GET.get("estado", "").strip()
    critico = request.GET.get("critico", "").strip()
    include_bajas = request.GET.get("include_bajas") == "1"
    if include_bajas and not estado:
        estado = ""
    equipos = (
        Equipo.objects.select_related(
            "centro_costo__division__sociedad",
            "marca",
            "sistema_operativo",
            "tipo_equipo",
            "modelo",
        )
        .order_by("-actualizado_en", "identificador")
    )
    if estado == "activo":
        equipos = equipos.filter(is_baja=False)
    elif estado == "baja":
        equipos = equipos.filter(is_baja=True)
    elif not include_bajas:
        equipos = equipos.filter(is_baja=False)
    if sociedad_id:
        equipos = equipos.filter(centro_costo__division__sociedad_id=sociedad_id)
    if division_id:
        equipos = equipos.filter(centro_costo__division_id=division_id)
    if centro_costo_id:
        equipos = equipos.filter(centro_costo_id=centro_costo_id)
    if marca_id:
        equipos = equipos.filter(marca_id=marca_id)
    if sistema_operativo_id:
        equipos = equipos.filter(sistema_operativo_id=sistema_operativo_id)
    if tipo_equipo_id:
        equipos = equipos.filter(tipo_equipo_id=tipo_equipo_id)
    if entidad:
        equipos = equipos.filter(entidad=entidad)
    if municipio:
        equipos = equipos.filter(municipio=municipio)
    if critico == "1":
        equipos = equipos.filter(infraestructura_critica=True)
    elif critico == "0":
        equipos = equipos.filter(infraestructura_critica=False)
    if texto:
        equipos = equipos.filter(
            Q(identificador__icontains=texto)
            | Q(clave__icontains=texto)
            | Q(numero_inventario__icontains=texto)
            | Q(numero_serie__icontains=texto)
            | Q(direccion_ip__icontains=texto)
            | Q(direccion_mac__icontains=texto)
            | Q(rpe_responsable__icontains=texto)
            | Q(nombre_responsable__icontains=texto)
            | Q(modelo__nombre__icontains=texto)
            | Q(marca__nombre__icontains=texto)
            | Q(nombre__icontains=texto)
        )

    total_encontrados = equipos.count()
    filtros_activos = any(
        [
            texto,
            sociedad_id,
            division_id,
            centro_costo_id,
            marca_id,
            sistema_operativo_id,
            tipo_equipo_id,
            entidad,
            municipio,
            estado,
            critico,
            include_bajas,
        ]
    )

    if request.GET.get("export") == "xlsx":
        if not can_edit(request.user):
            return render(request, "403.html", status=403)
        return _export_equipos_xlsx(equipos)

    paginator = Paginator(equipos, 25)
    page_obj = paginator.get_page(request.GET.get("page"))

    context = {
        "equipos": page_obj,
        "page_obj": page_obj,
        "sociedades": Equipo.objects.values_list(
            "centro_costo__division__sociedad__id",
            "centro_costo__division__sociedad__codigo",
            "centro_costo__division__sociedad__nombre",
        )
        .distinct()
        .order_by("centro_costo__division__sociedad__codigo"),
        "divisiones": Equipo.objects.values_list(
            "centro_costo__division__id",
            "centro_costo__division__codigo",
            "centro_costo__division__nombre",
        )
        .distinct()
        .order_by("centro_costo__division__codigo"),
        "centros_costo": Equipo.objects.values_list(
            "centro_costo__id",
            "centro_costo__codigo",
            "centro_costo__nombre",
        )
        .distinct()
        .order_by("centro_costo__codigo"),
        "marcas": Equipo.objects.values_list("marca__id", "marca__nombre")
        .distinct()
        .order_by("marca__nombre"),
        "sistemas_operativos": Equipo.objects.values_list(
            "sistema_operativo__id",
            "sistema_operativo__nombre",
        )
        .distinct()
        .order_by("sistema_operativo__nombre"),
        "tipos_equipo": Equipo.objects.values_list("tipo_equipo__id", "tipo_equipo__nombre")
        .distinct()
        .order_by("tipo_equipo__nombre"),
        "entidades": (
            Equipo.objects.exclude(entidad__isnull=True)
            .exclude(entidad__exact="")
            .values_list("entidad", flat=True)
            .distinct()
            .order_by("entidad")
        ),
        "municipios": (
            Equipo.objects.exclude(municipio__isnull=True)
            .exclude(municipio__exact="")
            .values_list("municipio", flat=True)
            .distinct()
            .order_by("municipio")
        ),
        "filtros": {
            "texto": texto,
            "sociedad": sociedad_id or "",
            "division": division_id or "",
            "centro_costo": centro_costo_id or "",
            "marca": marca_id or "",
            "sistema_operativo": sistema_operativo_id or "",
            "tipo_equipo": tipo_equipo_id or "",
            "entidad": entidad,
            "municipio": municipio,
            "estado": estado,
            "critico": critico,
        },
        "total_encontrados": total_encontrados,
        "filtros_activos": filtros_activos,
        "pagination_query": _build_querystring(request, exclude={"page"}),
        "export_query": _build_querystring(
            request, exclude={"page"}, extra={"export": "xlsx"}
        ),
        "can_baja": can_baja(request.user),
        "can_edit": can_edit(request.user),
        "can_import": can_import(request.user),
        "can_export": can_edit(request.user),
    }
    return render(request, "equipos/list.html", context)


@login_required
def equipo_detail(request, pk):
    bajas_queryset = BajaEquipo.objects.select_related("motivo", "usuario").order_by(
        "-fecha_baja"
    )
    equipo = get_object_or_404(
        Equipo.objects.select_related(
            "centro_costo__division__sociedad",
            "marca",
            "sistema_operativo",
            "tipo_equipo",
            "modelo",
        ).prefetch_related(Prefetch("bajas", queryset=bajas_queryset)),
        pk=pk,
    )
    context = {
        "equipo": equipo,
        "bajas": equipo.bajas.all(),
        "can_baja": can_baja(request.user),
        "can_edit": can_edit(request.user),
    }
    return render(request, "equipos/detail.html", context)


@login_required
def equipo_baja(request, pk):
    if not can_baja(request.user):
        return render(request, "403.html", status=403)

    equipo = get_object_or_404(Equipo, pk=pk)
    if equipo.is_baja:
        messages.warning(request, "El equipo ya se encuentra dado de baja.")
        return redirect("equipo_detail", pk=equipo.pk)

    motivos = MotivoBaja.objects.order_by("nombre")
    form_data = {
        "tipo_baja": "",
        "motivo": "",
        "comentarios": "",
    }

    if request.method == "POST":
        tipo_baja = request.POST.get("tipo_baja")
        motivo_id = request.POST.get("motivo")
        comentarios = request.POST.get("comentarios", "").strip()
        form_data.update(
            {
                "tipo_baja": tipo_baja or "",
                "motivo": motivo_id or "",
                "comentarios": comentarios,
            }
        )
        valid_choices = {choice[0] for choice in BajaEquipo.TipoBaja.choices}
        motivo = MotivoBaja.objects.filter(pk=motivo_id).first() if motivo_id else None
        if tipo_baja not in valid_choices:
            messages.error(request, "Seleccione un tipo de baja válido.")
        elif not motivo:
            messages.error(request, "Seleccione un motivo de baja válido.")
        else:
            equipo.is_baja = True
            equipo.fecha_baja = timezone.now()
            equipo.save(update_fields=["is_baja", "fecha_baja"])
            BajaEquipo.objects.create(
                equipo=equipo,
                fecha_baja=equipo.fecha_baja,
                tipo_baja=tipo_baja,
                motivo=motivo,
                comentarios=comentarios,
                usuario=request.user,
            )
            AuditLog.objects.create(
                usuario=request.user,
                accion="BAJA",
                resumen=(
                    f"Baja registrada para el equipo {equipo.identificador} "
                    f"({equipo.numero_serie}). Motivo: {motivo.nombre}."
                ),
                equipo=equipo,
            )
            messages.success(request, "La baja se registró correctamente.")
            return redirect("equipo_detail", pk=equipo.pk)

    context = {
        "equipo": equipo,
        "tipos_baja": BajaEquipo.TipoBaja.choices,
        "motivos": motivos,
        "form_data": form_data,
    }
    return render(request, "equipos/baja_form.html", context)


@login_required
def bajas_list(request):
    if not can_baja(request.user):
        return render(request, "403.html", status=403)

    bajas = (
        BajaEquipo.objects.select_related("equipo", "motivo", "usuario")
        .order_by("-fecha_baja")
    )
    fecha_desde = parse_date(request.GET.get("fecha_desde") or "")
    fecha_hasta = parse_date(request.GET.get("fecha_hasta") or "")
    motivo_id = request.GET.get("motivo")
    tipo_baja = request.GET.get("tipo_baja")
    if fecha_desde:
        bajas = bajas.filter(fecha_baja__date__gte=fecha_desde)
    if fecha_hasta:
        bajas = bajas.filter(fecha_baja__date__lte=fecha_hasta)
    if motivo_id:
        bajas = bajas.filter(motivo_id=motivo_id)
    if tipo_baja:
        bajas = bajas.filter(tipo_baja=tipo_baja)

    if request.GET.get("export") == "1":
        return _export_bajas_csv(bajas)

    context = {
        "bajas": bajas,
        "motivos": MotivoBaja.objects.order_by("nombre"),
        "tipos_baja": BajaEquipo.TipoBaja.choices,
        "filtros": {
            "fecha_desde": request.GET.get("fecha_desde") or "",
            "fecha_hasta": request.GET.get("fecha_hasta") or "",
            "motivo": motivo_id or "",
            "tipo_baja": tipo_baja or "",
        },
        "export_query": _build_querystring(request, extra={"export": "1"}),
    }
    return render(request, "bajas/list.html", context)


@login_required
def reportes_home(request):
    if not can_view_report(request.user):
        return render(request, "403.html", status=403)

    context = {
        "can_export": can_view_report(request.user),
    }
    return render(request, "reportes/index.html", context)


@login_required
def reporte_inventario_activo(request):
    if not can_view_report(request.user):
        return render(request, "403.html", status=403)

    equipos = (
        Equipo.objects.select_related(
            "centro_costo__division__sociedad",
            "marca",
            "sistema_operativo",
            "tipo_equipo",
            "modelo",
        )
        .filter(activo=True, is_baja=False)
        .order_by("identificador")
    )

    sociedad_id = request.GET.get("sociedad")
    division_id = request.GET.get("division")
    centro_costo_id = request.GET.get("centro_costo")
    marca_id = request.GET.get("marca")
    sistema_operativo_id = request.GET.get("sistema_operativo")
    tipo_equipo_id = request.GET.get("tipo_equipo")
    texto = request.GET.get("texto")

    if sociedad_id:
        equipos = equipos.filter(centro_costo__division__sociedad_id=sociedad_id)
    if division_id:
        equipos = equipos.filter(centro_costo__division_id=division_id)
    if centro_costo_id:
        equipos = equipos.filter(centro_costo_id=centro_costo_id)
    if marca_id:
        equipos = equipos.filter(marca_id=marca_id)
    if sistema_operativo_id:
        equipos = equipos.filter(sistema_operativo_id=sistema_operativo_id)
    if tipo_equipo_id:
        equipos = equipos.filter(tipo_equipo_id=tipo_equipo_id)
    if texto:
        equipos = equipos.filter(
            Q(identificador__icontains=texto)
            | Q(numero_inventario__icontains=texto)
            | Q(numero_serie__icontains=texto)
            | Q(nombre__icontains=texto)
        )

    if request.GET.get("export") == "1":
        if not can_view_report(request.user):
            return render(request, "403.html", status=403)
        return _export_inventario_activo_csv(equipos)

    paginator = Paginator(equipos, 25)
    page_obj = paginator.get_page(request.GET.get("page"))

    context = {
        "page_obj": page_obj,
        "sociedades": Equipo.objects.values_list(
            "centro_costo__division__sociedad__id",
            "centro_costo__division__sociedad__codigo",
            "centro_costo__division__sociedad__nombre",
        )
        .distinct()
        .order_by("centro_costo__division__sociedad__codigo"),
        "divisiones": Equipo.objects.values_list(
            "centro_costo__division__id",
            "centro_costo__division__codigo",
            "centro_costo__division__nombre",
        )
        .distinct()
        .order_by("centro_costo__division__codigo"),
        "centros_costo": Equipo.objects.values_list(
            "centro_costo__id",
            "centro_costo__codigo",
            "centro_costo__nombre",
        )
        .distinct()
        .order_by("centro_costo__codigo"),
        "marcas": Equipo.objects.values_list("marca__id", "marca__nombre")
        .distinct()
        .order_by("marca__nombre"),
        "sistemas_operativos": Equipo.objects.values_list(
            "sistema_operativo__id",
            "sistema_operativo__nombre",
        )
        .distinct()
        .order_by("sistema_operativo__nombre"),
        "tipos_equipo": Equipo.objects.values_list("tipo_equipo__id", "tipo_equipo__nombre")
        .distinct()
        .order_by("tipo_equipo__nombre"),
        "filtros": {
            "sociedad": sociedad_id or "",
            "division": division_id or "",
            "centro_costo": centro_costo_id or "",
            "marca": marca_id or "",
            "sistema_operativo": sistema_operativo_id or "",
            "tipo_equipo": tipo_equipo_id or "",
            "texto": texto or "",
        },
        "pagination_query": _build_querystring(request, exclude={"page"}),
        "export_query": _build_querystring(request, exclude={"page"}, extra={"export": "1"}),
        "can_export": can_view_report(request.user),
    }
    return render(request, "reportes/inventario_activo.html", context)


@login_required
def reporte_equipos_baja(request):
    if not can_view_report(request.user):
        return render(request, "403.html", status=403)

    bajas_queryset = BajaEquipo.objects.order_by("-fecha_baja")
    equipos = (
        Equipo.objects.select_related(
            "centro_costo__division__sociedad",
            "marca",
            "tipo_equipo",
        )
        .prefetch_related(Prefetch("bajas", queryset=bajas_queryset))
        .filter(is_baja=True)
        .order_by("-fecha_baja", "identificador")
    )

    fecha_desde = parse_date(request.GET.get("fecha_desde") or "")
    fecha_hasta = parse_date(request.GET.get("fecha_hasta") or "")
    motivo = request.GET.get("motivo")

    if fecha_desde:
        equipos = equipos.filter(fecha_baja__date__gte=fecha_desde)
    if fecha_hasta:
        equipos = equipos.filter(fecha_baja__date__lte=fecha_hasta)
    if motivo:
        equipos = equipos.filter(bajas__tipo_baja=motivo)

    equipos = equipos.distinct()

    if request.GET.get("export") == "1":
        if not can_view_report(request.user):
            return render(request, "403.html", status=403)
        return _export_equipos_baja_csv(equipos)

    context = {
        "equipos": equipos,
        "filtros": {
            "fecha_desde": request.GET.get("fecha_desde") or "",
            "fecha_hasta": request.GET.get("fecha_hasta") or "",
            "motivo": motivo or "",
        },
        "tipos_baja": BajaEquipo.TipoBaja.choices,
        "export_query": _build_querystring(request, extra={"export": "1"}),
        "can_export": can_view_report(request.user),
    }
    return render(request, "reportes/equipos_baja.html", context)


@login_required
def reporte_centro_costo(request):
    if not can_view_report(request.user):
        return render(request, "403.html", status=403)

    equipos = (
        Equipo.objects.select_related("centro_costo__division__sociedad")
    )

    sociedad_id = request.GET.get("sociedad")
    division_id = request.GET.get("division")

    if sociedad_id:
        equipos = equipos.filter(centro_costo__division__sociedad_id=sociedad_id)
    if division_id:
        equipos = equipos.filter(centro_costo__division_id=division_id)

    resumen = (
        equipos.values(
            "centro_costo__division__sociedad__codigo",
            "centro_costo__division__sociedad__nombre",
            "centro_costo__division__codigo",
            "centro_costo__division__nombre",
            "centro_costo__codigo",
            "centro_costo__nombre",
        )
        .annotate(
            total=Count("id"),
            total_activos=Count("id", filter=Q(activo=True, is_baja=False)),
            total_bajas=Count("id", filter=Q(is_baja=True)),
        )
        .order_by(
            "centro_costo__division__sociedad__codigo",
            "centro_costo__division__codigo",
            "centro_costo__codigo",
        )
    )

    sociedades = (
        Equipo.objects.values_list(
            "centro_costo__division__sociedad__id",
            "centro_costo__division__sociedad__codigo",
            "centro_costo__division__sociedad__nombre",
        )
        .distinct()
        .order_by("centro_costo__division__sociedad__codigo")
    )
    divisiones = (
        Equipo.objects.values_list(
            "centro_costo__division__id",
            "centro_costo__division__codigo",
            "centro_costo__division__nombre",
        )
        .distinct()
        .order_by("centro_costo__division__codigo")
    )

    resumen_agrupado = []
    sociedad_actual = None
    division_actual = None
    sociedad_total = 0
    division_total = 0
    sociedad_activos = 0
    sociedad_bajas = 0
    division_activos = 0
    division_bajas = 0

    for fila in resumen:
        total_activos = fila["total_activos"]
        total_bajas = fila["total_bajas"]
        total_general = fila["total"]
        pct_activos = (total_activos / total_general * 100) if total_general else 0
        pct_bajas = (total_bajas / total_general * 100) if total_general else 0
        sociedad_key = (
            fila["centro_costo__division__sociedad__codigo"],
            fila["centro_costo__division__sociedad__nombre"],
        )
        division_key = (
            fila["centro_costo__division__codigo"],
            fila["centro_costo__division__nombre"],
        )

        if sociedad_actual and sociedad_key != sociedad_actual:
            resumen_agrupado.append(
                {
                    "tipo": "division_subtotal",
                    "total": division_total,
                    "total_activos": division_activos,
                    "total_bajas": division_bajas,
                    "pct_activos": (division_activos / division_total * 100)
                    if division_total
                    else 0,
                    "pct_bajas": (division_bajas / division_total * 100)
                    if division_total
                    else 0,
                }
            )
            resumen_agrupado.append(
                {
                    "tipo": "sociedad_subtotal",
                    "total": sociedad_total,
                    "total_activos": sociedad_activos,
                    "total_bajas": sociedad_bajas,
                    "pct_activos": (sociedad_activos / sociedad_total * 100)
                    if sociedad_total
                    else 0,
                    "pct_bajas": (sociedad_bajas / sociedad_total * 100)
                    if sociedad_total
                    else 0,
                }
            )
            division_total = 0
            division_activos = 0
            division_bajas = 0
            sociedad_total = 0
            sociedad_activos = 0
            sociedad_bajas = 0
            division_actual = None

        if division_actual and division_key != division_actual:
            resumen_agrupado.append(
                {
                    "tipo": "division_subtotal",
                    "total": division_total,
                    "total_activos": division_activos,
                    "total_bajas": division_bajas,
                    "pct_activos": (division_activos / division_total * 100)
                    if division_total
                    else 0,
                    "pct_bajas": (division_bajas / division_total * 100)
                    if division_total
                    else 0,
                }
            )
            division_total = 0
            division_activos = 0
            division_bajas = 0

        if sociedad_key != sociedad_actual:
            resumen_agrupado.append(
                {
                    "tipo": "sociedad_header",
                    "sociedad_codigo": sociedad_key[0],
                    "sociedad_nombre": sociedad_key[1],
                }
            )
            sociedad_actual = sociedad_key
            division_actual = None

        if division_key != division_actual:
            resumen_agrupado.append(
                {
                    "tipo": "division_header",
                    "division_codigo": division_key[0],
                    "division_nombre": division_key[1],
                }
            )
            division_actual = division_key

        resumen_agrupado.append(
            {
                "tipo": "centro_costo",
                "centro_codigo": fila["centro_costo__codigo"],
                "centro_nombre": fila["centro_costo__nombre"],
                "total": total_general,
                "total_activos": total_activos,
                "total_bajas": total_bajas,
                "pct_activos": pct_activos,
                "pct_bajas": pct_bajas,
            }
        )
        division_total += total_general
        division_activos += total_activos
        division_bajas += total_bajas
        sociedad_total += total_general
        sociedad_activos += total_activos
        sociedad_bajas += total_bajas

    if sociedad_actual:
        resumen_agrupado.append(
            {
                "tipo": "division_subtotal",
                "total": division_total,
                "total_activos": division_activos,
                "total_bajas": division_bajas,
                "pct_activos": (division_activos / division_total * 100)
                if division_total
                else 0,
                "pct_bajas": (division_bajas / division_total * 100)
                if division_total
                else 0,
            }
        )
        resumen_agrupado.append(
            {
                "tipo": "sociedad_subtotal",
                "total": sociedad_total,
                "total_activos": sociedad_activos,
                "total_bajas": sociedad_bajas,
                "pct_activos": (sociedad_activos / sociedad_total * 100)
                if sociedad_total
                else 0,
                "pct_bajas": (sociedad_bajas / sociedad_total * 100)
                if sociedad_total
                else 0,
            }
        )

    export_type = request.GET.get("export")
    if export_type in {"1", "csv"}:
        if not can_view_report(request.user):
            return render(request, "403.html", status=403)
        return _export_centro_costo_csv(resumen)
    if export_type == "xlsx":
        if not can_view_report(request.user):
            return render(request, "403.html", status=403)
        return _export_centro_costo_xlsx(equipos)

    context = {
        "resumen": resumen_agrupado,
        "sociedades": sociedades,
        "divisiones": divisiones,
        "filtros": {
            "sociedad": sociedad_id or "",
            "division": division_id or "",
        },
        "export_csv_query": _build_querystring(request, extra={"export": "csv"}),
        "export_xlsx_query": _build_querystring(request, extra={"export": "xlsx"}),
        "can_export": can_view_report(request.user),
    }
    return render(request, "reportes/centro_costo.html", context)


@login_required
def reporte_responsables(request):
    if not can_view_report(request.user):
        return render(request, "403.html", status=403)

    equipos = (
        Equipo.objects.select_related("centro_costo__division__sociedad")
        .filter(activo=True, is_baja=False)
    )

    texto = request.GET.get("texto", "").strip()
    sociedad_id = request.GET.get("sociedad")
    division_id = request.GET.get("division")
    centro_costo_id = request.GET.get("centro_costo")

    if sociedad_id:
        equipos = equipos.filter(centro_costo__division__sociedad_id=sociedad_id)
    if division_id:
        equipos = equipos.filter(centro_costo__division_id=division_id)
    if centro_costo_id:
        equipos = equipos.filter(centro_costo_id=centro_costo_id)
    if texto:
        equipos = equipos.filter(
            Q(rpe_responsable__icontains=texto) | Q(nombre_responsable__icontains=texto)
        )

    resumen = (
        equipos.values(
            "rpe_responsable",
            "nombre_responsable",
            "centro_costo__division__sociedad__codigo",
            "centro_costo__division__sociedad__nombre",
            "centro_costo__division__codigo",
            "centro_costo__division__nombre",
            "centro_costo__codigo",
            "centro_costo__nombre",
        )
        .annotate(total=Count("id"))
        .order_by(
            "nombre_responsable",
            "rpe_responsable",
            "centro_costo__division__sociedad__codigo",
        )
    )

    if request.GET.get("export") == "1":
        if not can_view_report(request.user):
            return render(request, "403.html", status=403)
        return _export_responsables_csv(resumen)

    context = {
        "resumen": resumen,
        "sociedades": Equipo.objects.values_list(
            "centro_costo__division__sociedad__id",
            "centro_costo__division__sociedad__codigo",
            "centro_costo__division__sociedad__nombre",
        )
        .distinct()
        .order_by("centro_costo__division__sociedad__codigo"),
        "divisiones": Equipo.objects.values_list(
            "centro_costo__division__id",
            "centro_costo__division__codigo",
            "centro_costo__division__nombre",
        )
        .distinct()
        .order_by("centro_costo__division__codigo"),
        "centros_costo": Equipo.objects.values_list(
            "centro_costo__id",
            "centro_costo__codigo",
            "centro_costo__nombre",
        )
        .distinct()
        .order_by("centro_costo__codigo"),
        "filtros": {
            "sociedad": sociedad_id or "",
            "division": division_id or "",
            "centro_costo": centro_costo_id or "",
            "texto": texto,
        },
        "export_query": _build_querystring(request, extra={"export": "1"}),
        "can_export": can_view_report(request.user),
    }
    return render(request, "reportes/responsables.html", context)


@login_required
def reporte_resumen(request):
    if not can_view_report(request.user):
        return render(request, "403.html", status=403)

    activos = Equipo.objects.filter(activo=True, is_baja=False)
    bajas = Equipo.objects.filter(is_baja=True)

    resumen_sociedad = (
        activos.values(
            "centro_costo__division__sociedad__codigo",
            "centro_costo__division__sociedad__nombre",
        )
        .annotate(total=Count("id"))
        .order_by("centro_costo__division__sociedad__codigo")
    )
    resumen_tipo = (
        activos.values("tipo_equipo__nombre")
        .annotate(total=Count("id"))
        .order_by("tipo_equipo__nombre")
    )
    resumen_bajas = (
        BajaEquipo.objects.values("tipo_baja")
        .annotate(total=Count("id"))
        .order_by("tipo_baja")
    )
    motivo_lookup = dict(BajaEquipo.TipoBaja.choices)
    resumen_bajas = [
        {
            "tipo_baja": fila["tipo_baja"],
            "tipo_baja_display": motivo_lookup.get(fila["tipo_baja"], fila["tipo_baja"]),
            "total": fila["total"],
        }
        for fila in resumen_bajas
    ]

    context = {
        "total_equipos": Equipo.objects.count(),
        "total_activos": activos.count(),
        "total_bajas": bajas.count(),
        "resumen_sociedad": resumen_sociedad,
        "resumen_tipo": resumen_tipo,
        "resumen_bajas": resumen_bajas,
    }
    return render(request, "reportes/resumen.html", context)


@login_required
def equipo_editar(request, pk):
    if not can_edit(request.user):
        return render(request, "403.html", status=403)

    equipo = get_object_or_404(
        Equipo.objects.select_related(
            "centro_costo__division__sociedad",
            "marca",
            "sistema_operativo",
            "tipo_equipo",
            "modelo",
        ),
        pk=pk,
    )

    marcas = Marca.objects.values_list("id", "nombre").order_by("nombre")
    sistemas_operativos = SistemaOperativo.objects.values_list("id", "nombre").order_by("nombre")
    tipos_equipo = TipoEquipo.objects.values_list("id", "nombre").order_by("nombre")
    modelos = ModeloEquipo.objects.values_list("id", "nombre").order_by("nombre")
    centros_costo = CentroCosto.objects.values_list("id", "codigo", "nombre").order_by("codigo")

    form_data = {
        "domicilio": equipo.domicilio or "",
        "codigo_postal": equipo.codigo_postal or "",
        "antiguedad": equipo.antiguedad or "",
        "rpe_responsable": equipo.rpe_responsable or "",
        "nombre_responsable": equipo.nombre_responsable or "",
        "direccion_ip": equipo.direccion_ip or "",
        "direccion_mac": equipo.direccion_mac or "",
        "sistema_operativo": str(equipo.sistema_operativo_id or ""),
        "marca": str(equipo.marca_id or ""),
        "modelo": str(equipo.modelo_id or ""),
        "tipo_equipo": str(equipo.tipo_equipo_id or ""),
        "centro_costo": str(equipo.centro_costo_id or ""),
        "municipio": equipo.municipio or "",
        "entidad": equipo.entidad or "",
    }

    if request.method == "POST":
        form_data = {
            "domicilio": request.POST.get("domicilio", "").strip(),
            "codigo_postal": request.POST.get("codigo_postal", "").strip(),
            "antiguedad": request.POST.get("antiguedad", "").strip(),
            "rpe_responsable": request.POST.get("rpe_responsable", "").strip(),
            "nombre_responsable": request.POST.get("nombre_responsable", "").strip(),
            "direccion_ip": request.POST.get("direccion_ip", "").strip(),
            "direccion_mac": request.POST.get("direccion_mac", "").strip(),
            "sistema_operativo": request.POST.get("sistema_operativo", ""),
            "marca": request.POST.get("marca", ""),
            "modelo": request.POST.get("modelo", ""),
            "tipo_equipo": request.POST.get("tipo_equipo", ""),
            "centro_costo": request.POST.get("centro_costo", ""),
            "municipio": request.POST.get("municipio", "").strip(),
            "entidad": request.POST.get("entidad", "").strip(),
        }

        errors = []
        if form_data["direccion_ip"] and not IP_REGEX.match(form_data["direccion_ip"]):
            errors.append("La dirección IP no tiene un formato válido.")
        if form_data["direccion_mac"] and not MAC_REGEX.match(form_data["direccion_mac"]):
            errors.append("La dirección MAC no tiene un formato válido.")

        centro_costo = CentroCosto.objects.filter(pk=form_data["centro_costo"]).first()
        if not centro_costo:
            errors.append("Seleccione un centro de costo válido.")

        if form_data["marca"] and not Marca.objects.filter(pk=form_data["marca"]).exists():
            errors.append("Seleccione una marca válida.")
        if form_data["modelo"] and not ModeloEquipo.objects.filter(pk=form_data["modelo"]).exists():
            errors.append("Seleccione un modelo válido.")
        if form_data["sistema_operativo"] and not SistemaOperativo.objects.filter(
            pk=form_data["sistema_operativo"]
        ).exists():
            errors.append("Seleccione un sistema operativo válido.")
        if form_data["tipo_equipo"] and not TipoEquipo.objects.filter(
            pk=form_data["tipo_equipo"]
        ).exists():
            errors.append("Seleccione un tipo de equipo válido.")

        if errors:
            for error in errors:
                messages.error(request, error)
        else:
            original = {
                "marca": equipo.marca.nombre if equipo.marca else "",
                "modelo": equipo.modelo.nombre if equipo.modelo else "",
                "sistema_operativo": equipo.sistema_operativo.nombre
                if equipo.sistema_operativo
                else "",
                "tipo_equipo": equipo.tipo_equipo.nombre if equipo.tipo_equipo else "",
                "direccion_ip": equipo.direccion_ip or "",
                "direccion_mac": equipo.direccion_mac or "",
                "domicilio": equipo.domicilio or "",
                "codigo_postal": equipo.codigo_postal or "",
                "antiguedad": equipo.antiguedad or "",
                "rpe_responsable": equipo.rpe_responsable or "",
                "nombre_responsable": equipo.nombre_responsable or "",
                "centro_costo": (
                    f"{equipo.centro_costo.codigo} - {equipo.centro_costo.nombre}"
                ),
                "municipio": equipo.municipio or "",
                "entidad": equipo.entidad or "",
            }

            equipo.direccion_ip = form_data["direccion_ip"] or None
            equipo.direccion_mac = form_data["direccion_mac"] or None
            equipo.domicilio = form_data["domicilio"] or None
            equipo.codigo_postal = form_data["codigo_postal"] or None
            equipo.antiguedad = form_data["antiguedad"] or None
            equipo.rpe_responsable = form_data["rpe_responsable"] or None
            equipo.nombre_responsable = form_data["nombre_responsable"] or None
            equipo.centro_costo_id = form_data["centro_costo"]
            equipo.marca_id = form_data["marca"] or None
            equipo.modelo_id = form_data["modelo"] or None
            equipo.sistema_operativo_id = form_data["sistema_operativo"] or None
            equipo.tipo_equipo_id = form_data["tipo_equipo"] or None
            equipo.municipio = form_data["municipio"] or None
            equipo.entidad = form_data["entidad"] or None

            equipo.save()

            updated = {
                "marca": equipo.marca.nombre if equipo.marca else "",
                "modelo": equipo.modelo.nombre if equipo.modelo else "",
                "sistema_operativo": equipo.sistema_operativo.nombre
                if equipo.sistema_operativo
                else "",
                "tipo_equipo": equipo.tipo_equipo.nombre if equipo.tipo_equipo else "",
                "direccion_ip": equipo.direccion_ip or "",
                "direccion_mac": equipo.direccion_mac or "",
                "domicilio": equipo.domicilio or "",
                "codigo_postal": equipo.codigo_postal or "",
                "antiguedad": equipo.antiguedad or "",
                "rpe_responsable": equipo.rpe_responsable or "",
                "nombre_responsable": equipo.nombre_responsable or "",
                "centro_costo": (
                    f"{equipo.centro_costo.codigo} - {equipo.centro_costo.nombre}"
                ),
                "municipio": equipo.municipio or "",
                "entidad": equipo.entidad or "",
            }

            labels = {
                "marca": "Marca",
                "modelo": "Modelo",
                "sistema_operativo": "SO",
                "tipo_equipo": "Tipo",
                "direccion_ip": "IP",
                "direccion_mac": "MAC",
                "domicilio": "Domicilio",
                "codigo_postal": "CP",
                "antiguedad": "Antigüedad",
                "rpe_responsable": "RPE responsable",
                "nombre_responsable": "Nombre responsable",
                "centro_costo": "Centro de costo",
                "municipio": "Municipio",
                "entidad": "Entidad",
            }
            cambios = []
            for key, label in labels.items():
                if original[key] != updated[key]:
                    cambios.append(f"{label}: '{original[key]}' → '{updated[key]}'")
            resumen = (
                f"Equipo {equipo.identificador} actualizado. "
                + "; ".join(cambios)
                if cambios
                else f"Equipo {equipo.identificador} actualizado sin cambios detectados."
            )
            AuditLog.objects.create(
                usuario=request.user,
                accion="EDITAR_EQUIPO",
                resumen=resumen,
                equipo=equipo,
            )
            messages.success(request, "El equipo se actualizó correctamente.")
            return redirect("equipo_detail", pk=equipo.pk)

    context = {
        "equipo": equipo,
        "form_data": form_data,
        "marcas": marcas,
        "modelos": modelos,
        "sistemas_operativos": sistemas_operativos,
        "tipos_equipo": tipos_equipo,
        "centros_costo": centros_costo,
    }
    return render(request, "equipos/editar.html", context)


@login_required
def auditoria_list(request):
    if not can_audit(request.user):
        return render(request, "403.html", status=403)

    fecha_desde = parse_date(request.GET.get("fecha_desde") or "")
    fecha_hasta = parse_date(request.GET.get("fecha_hasta") or "")
    usuario_id = request.GET.get("usuario")
    accion = request.GET.get("accion")
    entidad = request.GET.get("entidad")

    audit_logs = AuditLog.objects.select_related("usuario", "equipo").order_by("-fecha")
    import_logs = ImportLog.objects.select_related("usuario").order_by("-fecha")

    if fecha_desde:
        audit_logs = audit_logs.filter(fecha__date__gte=fecha_desde)
        import_logs = import_logs.filter(fecha__date__gte=fecha_desde)
    if fecha_hasta:
        audit_logs = audit_logs.filter(fecha__date__lte=fecha_hasta)
        import_logs = import_logs.filter(fecha__date__lte=fecha_hasta)
    if usuario_id:
        audit_logs = audit_logs.filter(usuario_id=usuario_id)
        import_logs = import_logs.filter(usuario_id=usuario_id)

    if accion:
        audit_logs = audit_logs.filter(accion=accion)
        if accion != "IMPORT":
            import_logs = import_logs.none()
    if entidad:
        audit_logs = audit_logs.filter(equipo__entidad=entidad)
        import_logs = import_logs.none()

    registros = [
        {
            "fecha": log.fecha,
            "usuario": log.usuario,
            "accion": log.accion,
            "resumen": log.resumen,
            "equipo": log.equipo,
        }
        for log in audit_logs
    ]
    registros.extend(
        [
            {
                "fecha": log.fecha,
                "usuario": log.usuario,
                "accion": "IMPORT",
                "resumen": (
                    "Importación CSV ejecutada. "
                    f"Total: {log.total_filas}, "
                    f"Creados: {log.creados}, "
                    f"Actualizados: {log.actualizados}, "
                    f"Omitidos: {log.omitidos}, "
                    f"Errores: {log.errores}."
                ),
                "equipo": None,
            }
            for log in import_logs
        ]
    )
    registros.sort(key=lambda item: item["fecha"], reverse=True)

    if request.GET.get("export") == "1":
        return _export_auditoria_csv(registros)

    paginator = Paginator(registros, 25)
    page_obj = paginator.get_page(request.GET.get("page"))

    acciones_set = set(audit_logs.values_list("accion", flat=True))
    if import_logs.exists():
        acciones_set.add("IMPORT")
    acciones_disponibles = sorted(acciones_set)

    User = get_user_model()
    usuarios = (
        User.objects.filter(id__in=audit_logs.values_list("usuario_id", flat=True))
        | User.objects.filter(id__in=import_logs.values_list("usuario_id", flat=True))
    )
    usuarios = usuarios.distinct().order_by("username")

    context = {
        "registros": page_obj,
        "page_obj": page_obj,
        "usuarios": usuarios,
        "acciones": acciones_disponibles,
        "entidades": (
            Equipo.objects.exclude(entidad__isnull=True)
            .exclude(entidad__exact="")
            .values_list("entidad", flat=True)
            .distinct()
            .order_by("entidad")
        ),
        "filtros": {
            "fecha_desde": request.GET.get("fecha_desde") or "",
            "fecha_hasta": request.GET.get("fecha_hasta") or "",
            "usuario": usuario_id or "",
            "accion": accion or "",
            "entidad": entidad or "",
        },
        "export_query": _build_querystring(request, extra={"export": "1"}),
        "pagination_query": _build_querystring(request, exclude={"page"}),
    }
    return render(request, "auditoria/list.html", context)


def _export_inventario_activo_csv(equipos):
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="reporte_inventario_activo.csv"'
    writer = csv.writer(response)
    writer.writerow(
        [
            "Identificador",
            "Inventario",
            "Serie",
            "Nombre",
            "Sociedad",
            "Division",
            "Centro de costo",
            "Marca",
            "Sistema operativo",
            "Tipo equipo",
            "Modelo",
        ]
    )
    for equipo in equipos:
        writer.writerow(
            [
                equipo.identificador,
                equipo.numero_inventario,
                equipo.numero_serie,
                equipo.nombre,
                equipo.centro_costo.division.sociedad,
                equipo.centro_costo.division,
                equipo.centro_costo,
                equipo.marca or "",
                equipo.sistema_operativo or "",
                equipo.tipo_equipo or "",
                equipo.modelo or "",
            ]
        )
    return response


def _export_equipos_xlsx(equipos):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Equipos"

    headers = [
        "Nombre",
        "Identificador",
        "Inventario",
        "Serie",
        "Centro de costo",
        "Marca",
        "Sistema operativo",
        "RPE responsable",
        "Nombre responsable",
        "Municipio",
        "Domicilio",
        "Estado",
        "Crítico",
    ]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for equipo in equipos:
        sheet.append(
            [
                equipo.nombre,
                equipo.identificador,
                equipo.numero_inventario or "",
                equipo.numero_serie,
                str(equipo.centro_costo),
                str(equipo.marca or ""),
                str(equipo.sistema_operativo or ""),
                equipo.rpe_responsable or "",
                equipo.nombre_responsable or "",
                equipo.municipio or "",
                equipo.domicilio or "",
                "Baja" if equipo.is_baja else "Activo",
                "Sí" if equipo.infraestructura_critica else "No",
            ]
        )

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="equipos.xlsx"'
    workbook.save(response)
    return response


def _export_bajas_csv(bajas):
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="bajas_equipos.csv"'
    writer = csv.writer(response)
    writer.writerow(
        [
            "Identificador",
            "Inventario",
            "Serie",
            "Tipo de baja",
            "Motivo",
            "Fecha",
            "Usuario",
            "Comentarios",
        ]
    )
    for baja in bajas:
        writer.writerow(
            [
                baja.equipo.identificador,
                baja.equipo.numero_inventario,
                baja.equipo.numero_serie,
                baja.get_tipo_baja_display(),
                baja.motivo.nombre if baja.motivo else "",
                baja.fecha_baja.strftime("%Y-%m-%d %H:%M"),
                baja.usuario.get_username() if baja.usuario else "",
                baja.comentarios or "",
            ]
        )
    return response


def _export_equipos_baja_csv(equipos):
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="reporte_equipos_baja.csv"'
    writer = csv.writer(response)
    writer.writerow(
        [
            "Identificador",
            "Inventario",
            "Serie",
            "Nombre",
            "Fecha baja",
            "Motivo",
            "Sociedad",
            "Division",
            "Centro de costo",
            "Marca",
            "Tipo equipo",
        ]
    )
    for equipo in equipos:
        baja = equipo.bajas.all().first()
        writer.writerow(
            [
                equipo.identificador,
                equipo.numero_inventario,
                equipo.numero_serie,
                equipo.nombre,
                equipo.fecha_baja.strftime("%Y-%m-%d %H:%M") if equipo.fecha_baja else "",
                baja.get_tipo_baja_display() if baja else "",
                equipo.centro_costo.division.sociedad,
                equipo.centro_costo.division,
                equipo.centro_costo,
                equipo.marca or "",
                equipo.tipo_equipo or "",
            ]
        )
    return response


def _export_centro_costo_csv(resumen):
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="reporte_centro_costo.csv"'
    writer = csv.writer(response)
    writer.writerow(
        [
            "Sociedad",
            "Division",
            "Centro de costo",
            "Total equipos",
            "% activos",
            "% bajas",
        ]
    )
    for fila in resumen:
        total = fila["total"]
        activos = fila["total_activos"]
        bajas = fila["total_bajas"]
        pct_activos = round((activos / total * 100), 1) if total else 0
        pct_bajas = round((bajas / total * 100), 1) if total else 0
        writer.writerow(
            [
                f"{fila['centro_costo__division__sociedad__codigo']} - "
                f"{fila['centro_costo__division__sociedad__nombre']}",
                f"{fila['centro_costo__division__codigo']} - "
                f"{fila['centro_costo__division__nombre']}",
                f"{fila['centro_costo__codigo']} - {fila['centro_costo__nombre']}",
                total,
                f"{pct_activos}%",
                f"{pct_bajas}%",
            ]
        )
    return response


def _autosize_columns(worksheet):
    for column_cells in worksheet.columns:
        length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            if cell.value is not None:
                length = max(length, len(str(cell.value)))
        worksheet.column_dimensions[column].width = min(length + 2, 60)


def _export_centro_costo_xlsx(equipos):
    workbook = Workbook()
    resumen_sheet = workbook.active
    resumen_sheet.title = "Resumen"

    totales = equipos.aggregate(
        total_activos=Count("id", filter=Q(activo=True, is_baja=False)),
        total_bajas=Count("id", filter=Q(is_baja=True)),
        total_general=Count("id"),
    )
    total_general = totales["total_general"] or 0
    total_activos_count = totales["total_activos"] or 0
    total_bajas_count = totales["total_bajas"] or 0
    pct_activos = round((total_activos_count / total_general * 100), 1) if total_general else 0
    pct_bajas = round((total_bajas_count / total_general * 100), 1) if total_general else 0

    resumen_headers = ["Total general", "Activos", "% activos", "Bajas", "% bajas"]
    resumen_sheet.append(resumen_headers)
    resumen_sheet.append(
        [total_general, total_activos_count, f"{pct_activos}%", total_bajas_count, f"{pct_bajas}%"]
    )
    for cell in resumen_sheet[1]:
        cell.font = Font(bold=True)
    _autosize_columns(resumen_sheet)

    detalle_sheet = workbook.create_sheet(title="Detalle")
    detalle_headers = [
        "Identificador",
        "Inventario",
        "Serie",
        "Nombre",
        "Sociedad",
        "División",
        "Centro de costo",
        "Estado",
    ]
    detalle_sheet.append(detalle_headers)
    for cell in detalle_sheet[1]:
        cell.font = Font(bold=True)
    for equipo in equipos.select_related("centro_costo__division__sociedad"):
        estado = "Baja" if equipo.is_baja else "Activo"
        detalle_sheet.append(
            [
                equipo.identificador,
                equipo.numero_inventario,
                equipo.numero_serie,
                equipo.nombre,
                equipo.centro_costo.division.sociedad,
                equipo.centro_costo.division,
                equipo.centro_costo,
                estado,
            ]
        )
    _autosize_columns(detalle_sheet)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="reporte_centro_costo.xlsx"'
    workbook.save(response)
    return response


def _export_responsables_csv(resumen):
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="reporte_responsables.csv"'
    writer = csv.writer(response)
    writer.writerow(
        [
            "Responsable",
            "RPE",
            "Sociedad",
            "Division",
            "Centro de costo",
            "Total equipos",
        ]
    )
    for fila in resumen:
        writer.writerow(
            [
                fila["nombre_responsable"] or "",
                fila["rpe_responsable"] or "",
                f"{fila['centro_costo__division__sociedad__codigo']} - "
                f"{fila['centro_costo__division__sociedad__nombre']}",
                f"{fila['centro_costo__division__codigo']} - "
                f"{fila['centro_costo__division__nombre']}",
                f"{fila['centro_costo__codigo']} - {fila['centro_costo__nombre']}",
                fila["total"],
            ]
        )
    return response


def _export_auditoria_csv(registros):
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="auditoria.csv"'
    writer = csv.writer(response)
    writer.writerow(["Fecha", "Usuario", "Accion", "Resumen", "Equipo"])
    for registro in registros:
        writer.writerow(
            [
                registro["fecha"].strftime("%Y-%m-%d %H:%M"),
                registro["usuario"].get_username() if registro["usuario"] else "",
                registro["accion"],
                registro["resumen"],
                registro["equipo"].identificador if registro["equipo"] else "",
            ]
        )
    return response
